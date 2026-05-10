import openpyxl
import curl_cffi
import json
from random import uniform
from math import ceil
from re import sub
from time import sleep
from utils import clean_spreadsheet, fetch_with_backoff, get_random_user_agent, get_xlsx_filepath


def page_payload(page: int) -> str:
    payload = {
        "FilteringOptions": {
            "undefined": 0,
            "OngoingCharge": {

            },
            "RangeId": None,
            "RangeName": "",
            "CategoryId": None,
            "Category2Id": None,
            "PriipProductCode": None,
            "DefaultCategoryId": None,
            "DefaultCategory2Id": None,
            "ForSaleIn": None,
            "ShowMainUnits": False,
            "MPCategoryCode": None
        },
        "ProjectName": "quilter",
        "LanguageCode": "en-gb",
        "LanguageId": "1",
        "Theme": "quilter_new",
        "SortingStyle": "1",
        "PageNo": page,  # THIS
        "PageSize": 30,
        "OrderBy": "UnitName:init",
        "IsAscOrder": True,
        "OverrideDocumentCountryCode": None,
        "ToolId": "1",
        "PrefetchPages": 67,
        "PrefetchPageStart": page,  # THIS
        "OverridenThemeName": "quilter_new",
        "ForSaleIn": "",
        "ValidateFeResearchAccess": False,
        "HasFeResearchFullAccess": False,
        "EnableSedolSearch": "false",
        "GrsProjectId": "17200144",
        "ShowMainUnitExpansion": False,
        "UseCombinedOngoingChargeTER": False
    }
    return json.dumps(payload)


def quilter_runner() -> None:
    out_xlsx = get_xlsx_filepath("quilter.xlsx")
    clean_spreadsheet(out_xlsx)
    BASE_URL = 'https://digitalfundservice.feprecisionplus.com/FundDataService.svc/GetRowIdList?jsonString='
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'en-GB,en;q=0.5',
        'Origin': 'https://digital.feprecisionplus.com',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Referer': 'https://digital.feprecisionplus.com/',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-site',
        'Content-Type': 'application/json'
    }
    headers.update(get_random_user_agent())
    page = 1
    wb = openpyxl.load_workbook(out_xlsx)
    ws = wb["Funds"]
    sheet_current_row = 2

    url = "https://digital.feprecisionplus.com"
    has_next = True
    while has_next:
        payload_encoded = page_payload(page)
        res = fetch_with_backoff(url=BASE_URL + payload_encoded,
                                 headers=headers)
        json_data = res.json()
       # if page == 1:
        total_funds = ceil(json_data["TotalRows"] / 30)
        print(
            f'[Quilter] Total Funds found: {json_data["TotalRows"]} [###]')
        data = sub(r'\r\n\s', '', json_data["Units"])
        funds_per_page = json.loads(data)["DataList"]
        print(f'[#] Page {page} of {total_funds}')
        for fund in funds_per_page:
            # fund_url = f'{url}{fund["FundInfo"]["ResponsiveFactsheetLink"]}'
            f = fund.get("FundInfo")
            if f:
                fund_url = f.get("ResponsiveFactsheetLink")
                name = f.get("Name")
                isin = f.get("ISIN")
                if name:
                    ws.cell(sheet_current_row, 1).value = name
                if isin:
                    ws.cell(sheet_current_row,
                            2).value = isin
                if fund_url:
                    fund_url = f'{url}{fund_url}'
                    c = ws.cell(sheet_current_row, 3, fund_url)
                    c.hyperlink = fund_url
                    c.style = "Hyperlink"
            sheet_current_row += 1
        has_next = False if page == total_funds else True
        wb.save(out_xlsx)
        page += 1
        sleep(uniform(0.2, 0.5))
    wb.close()
